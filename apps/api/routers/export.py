"""Task result export — download completed tasks as CSV, JSON, or Excel."""
from __future__ import annotations
import csv
import io
import json
from datetime import datetime, timezone
from typing import Optional
from uuid import UUID

import structlog
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_

from core.auth import get_current_user_id
from core.database import get_db
from models.db import TaskDB, TaskAssignmentDB, OrganizationDB, OrgMemberDB

logger = structlog.get_logger()
router = APIRouter(prefix="/v1/tasks/export", tags=["export"])

VALID_FORMATS = ("csv", "json", "xlsx")


def _summarise(data: Optional[dict], max_len: int = 200) -> Optional[str]:
    """Convert a dict to a short summary string."""
    if data is None:
        return None
    try:
        s = json.dumps(data, default=str)
        return s if len(s) <= max_len else s[:max_len] + "…"
    except (TypeError, ValueError, OverflowError):
        return str(data)[:max_len]


def _fmt_dt(dt: Optional[datetime]) -> Optional[str]:
    if dt is None:
        return None
    return dt.isoformat()


def _build_xlsx(rows: list[dict]) -> bytes:
    """Build an Excel workbook from export rows and return bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    if not rows:
        ws.append(["No tasks found"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Build flat rows (exclude submissions list)
    flat_rows = []
    for r in rows:
        flat = {k: v for k, v in r.items() if k != "submissions"}
        # Flatten lists / dicts to JSON strings
        for k, v in flat.items():
            if isinstance(v, (dict, list)):
                flat[k] = json.dumps(v, default=str)
        flat_rows.append(flat)

    headers = list(flat_rows[0].keys())

    # Header row styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(flat_rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            val = row.get(header)
            # Convert None to empty string
            ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")
        # Alternating row color
        if row_idx % 2 == 0:
            row_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill

    # Auto-fit columns (approximate)
    for col_idx, header in enumerate(headers, start=1):
        max_len = max(
            len(str(header)),
            max((len(str(row.get(header) or "")) for row in flat_rows), default=0),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add a Summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total tasks", len(rows)])
    status_counts: dict[str, int] = {}
    for r in rows:
        s = r.get("status") or "unknown"
        status_counts[s] = status_counts.get(s, 0) + 1
    for status, count in sorted(status_counts.items()):
        ws_summary.append([f"Status: {status}", count])
    ws_summary.append(["Exported at", datetime.now(timezone.utc).isoformat()])

    for cell in ws_summary["A"]:
        cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("")
async def export_tasks(
    format: str = Query("csv", description="Export format: csv | json | xlsx"),
    status: Optional[str] = Query(
        None,
        description="Filter by status: completed | failed | all (default: all)",
    ),
    type: Optional[str] = Query(None, description="Filter by task type"),
    execution_mode: Optional[str] = Query(None, description="ai | human"),
    from_date: Optional[str] = Query(None, description="ISO date start: YYYY-MM-DD"),
    to_date: Optional[str] = Query(None, description="ISO date end: YYYY-MM-DD"),
    org_id: Optional[UUID] = Query(None, description="Export tasks for an org (must be member)"),
    include_submissions: bool = Query(
        False,
        description="Include worker submission data in export (human tasks only)",
    ),
    db: AsyncSession = Depends(get_db),
    user_id: str = Depends(get_current_user_id),
):
    """
    Download all your tasks as CSV, JSON, or Excel (.xlsx).
    Supports filtering by status, type, date range, and org.
    """
    if format not in VALID_FORMATS:
        raise HTTPException(status_code=400, detail=f"format must be one of: {', '.join(VALID_FORMATS)}")

    # Build query
    conditions = []

    if org_id:
        # Verify membership
        mem_result = await db.execute(
            select(OrgMemberDB).where(
                OrgMemberDB.org_id == org_id,
                OrgMemberDB.user_id == user_id,
            )
        )
        if not mem_result.scalar_one_or_none():
            raise HTTPException(status_code=403, detail="Not a member of this organization")
        conditions.append(TaskDB.org_id == org_id)
    else:
        conditions.append(TaskDB.user_id == user_id)

    if status and status != "all":
        conditions.append(TaskDB.status == status)

    if type:
        conditions.append(TaskDB.type == type)

    if execution_mode:
        conditions.append(TaskDB.execution_mode == execution_mode)

    if from_date:
        try:
            dt = datetime.fromisoformat(from_date).replace(tzinfo=timezone.utc)
            conditions.append(TaskDB.created_at >= dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid from_date format (use YYYY-MM-DD)")

    if to_date:
        try:
            dt = datetime.fromisoformat(to_date).replace(tzinfo=timezone.utc)
            # End of day
            dt = dt.replace(hour=23, minute=59, second=59)
            conditions.append(TaskDB.created_at <= dt)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid to_date format (use YYYY-MM-DD)")

    MAX_EXPORT_ROWS = 10_000
    q = select(TaskDB).where(and_(*conditions)).order_by(TaskDB.created_at.desc()).limit(MAX_EXPORT_ROWS)
    result = await db.execute(q)
    tasks = result.scalars().all()

    # Optionally load submissions for human tasks
    submissions_by_task: dict[str, list] = {}
    if include_submissions:
        task_ids = [t.id for t in tasks if t.execution_mode == "human"]
        if task_ids:
            asgn_result = await db.execute(
                select(TaskAssignmentDB).where(
                    TaskAssignmentDB.task_id.in_(task_ids),
                    TaskAssignmentDB.status.in_(["submitted", "approved", "rejected"]),
                )
            )
            for a in asgn_result.scalars().all():
                key = str(a.task_id)
                if key not in submissions_by_task:
                    submissions_by_task[key] = []
                submissions_by_task[key].append({
                    "assignment_id": str(a.id),
                    "worker_id": str(a.worker_id),
                    "status": a.status,
                    "response": a.response,
                    "worker_note": a.worker_note,
                    "earnings_credits": a.earnings_credits,
                    "submitted_at": _fmt_dt(a.submitted_at),
                })

    # Build export rows
    rows = []
    for t in tasks:
        row = {
            "task_id": str(t.id),
            "type": t.type,
            "status": t.status,
            "execution_mode": t.execution_mode,
            "priority": t.priority,
            "consensus_strategy": t.consensus_strategy,
            "dispute_status": t.dispute_status,
            "created_at": _fmt_dt(t.created_at),
            "started_at": _fmt_dt(t.started_at),
            "completed_at": _fmt_dt(t.completed_at),
            "credits_used": t.credits_used,
            "duration_ms": t.duration_ms,
            "assignments_required": t.assignments_required,
            "assignments_completed": t.assignments_completed,
            "worker_reward_credits": t.worker_reward_credits,
            "is_gold_standard": t.is_gold_standard,
            "org_id": str(t.org_id) if t.org_id else None,
            "input": json.dumps(t.input, default=str) if t.input else None,
            "output": json.dumps(t.output, default=str) if t.output else None,
            "error": t.error,
            "metadata": json.dumps(t.task_metadata, default=str) if t.task_metadata else None,
        }
        if include_submissions:
            row["submissions"] = submissions_by_task.get(str(t.id), [])
        rows.append(row)

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    if format == "json":
        payload = json.dumps(
            {"exported_at": datetime.now(timezone.utc).isoformat(), "count": len(rows), "tasks": rows},
            indent=2,
            default=str,
        )
        return StreamingResponse(
            iter([payload]),
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="tasks_export_{ts}.json"'},
        )

    if format == "xlsx":
        xlsx_bytes = _build_xlsx(rows)
        return StreamingResponse(
            iter([xlsx_bytes]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="tasks_export_{ts}.xlsx"'},
        )

    # CSV format
    if not rows:
        # Return empty CSV with headers
        fieldnames = [
            "task_id", "type", "status", "execution_mode", "priority",
            "consensus_strategy", "dispute_status",
            "created_at", "started_at", "completed_at",
            "credits_used", "duration_ms",
            "assignments_required", "assignments_completed", "worker_reward_credits",
            "is_gold_standard", "org_id", "input", "output", "error", "metadata",
        ]
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=fieldnames)
        writer.writeheader()
    else:
        fieldnames = [k for k in rows[0].keys() if k != "submissions"]
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="tasks_export_{ts}.csv"'},
    )
